#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import build_record, ensure_parent_dir, write_jsonl


TEXT_HINTS = ("text", "comment", "review", "content", "body", "message", "feedback", "note", "description")
TITLE_HINTS = ("title", "subject", "heading", "name")


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def choose_text(parts: dict[str, str]) -> str:
    prioritized = [value for key, value in parts.items() if any(hint in key.lower() for hint in TEXT_HINTS) and value]
    if prioritized:
        return "\n".join(prioritized)
    return "\n".join(value for value in parts.values() if value)


def choose_title(parts: dict[str, str], row_number: int) -> str:
    for key, value in parts.items():
        if any(hint in key.lower() for hint in TITLE_HINTS) and value:
            return value[:120]
    return f"Row {row_number}"


def parse_workbook(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    records: list[dict] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [normalize_cell(value) or f"column_{index}" for index, value in enumerate(rows[0], start=1)]
        for row_offset, row in enumerate(rows[1:], start=2):
            row_map = {headers[index]: normalize_cell(value) for index, value in enumerate(row)}
            text = choose_text(row_map)
            if not text:
                continue
            records.append(
                build_record(
                    source_file=str(input_path),
                    record_id=f"{sheet.title}-{row_offset}",
                    title=choose_title(row_map, row_offset),
                    text=text,
                    metadata={
                        "source_type": "excel",
                        "sheet_name": sheet.title,
                        "row_number": row_offset,
                        "column_map": row_map,
                    },
                )
            )
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse an Excel workbook into normalized jsonl records.")
    parser.add_argument("input", type=Path, help="Workbook path such as .xlsx")
    parser.add_argument("output", type=Path, help="Output jsonl path")
    args = parser.parse_args()

    records = parse_workbook(args.input)
    ensure_parent_dir(args.output)
    write_jsonl(args.output, records)
    print(f"Wrote {len(records)} record(s) to {args.output}")


if __name__ == "__main__":
    main()
